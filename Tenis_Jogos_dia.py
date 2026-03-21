import streamlit as st
import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="ATP/WTA Live Matches", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    [data-testid="stSidebar"] {
        display: none;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("# 🎾 ATP/WTA LIVE MATCHES ANALYZER")
st.markdown("Live tennis matches with competitive analysis (22-23 games)")
st.markdown("---")

# SCRAPE LIVE DATA
st.markdown("## 📥 LOADING LIVE MATCH DATA")

@st.cache_data(ttl=3600)
def scrape_flashscore():
    """Scrape live matches from FlashScore"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        # FlashScore URLs for ATP and WTA
        urls = {
            'ATP': 'https://www.flashscore.com/tennis/atp/',
            'WTA': 'https://www.flashscore.com/tennis/wta/'
        }
        
        all_matches = []
        
        for tour, url in urls.items():
            try:
                response = requests.get(url, headers=headers, timeout=10)
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Look for match data
                matches = soup.find_all('div', {'class': 'event__match'})
                
                if matches:
                    st.success(f"✅ Found {len(matches)} {tour} matches on FlashScore")
                    
                    for match in matches[:20]:  # Get top 20 matches
                        try:
                            # Extract match info
                            teams = match.find_all('span', {'class': 'event__participant'})
                            score = match.find('span', {'class': 'event__score'})
                            
                            if len(teams) >= 2 and score:
                                player1 = teams[0].text.strip()
                                player2 = teams[1].text.strip()
                                score_text = score.text.strip()
                                
                                all_matches.append({
                                    'Winner': player1,
                                    'Loser': player2,
                                    'Score': score_text,
                                    'Tour': tour,
                                    'Surface': 'Hard',
                                    'Date': datetime.now().strftime('%Y-%m-%d'),
                                    'W1': 0, 'L1': 0, 'W2': 0, 'L2': 0, 'W3': 0, 'L3': 0
                                })
                        except:
                            continue
                            
            except Exception as e:
                st.warning(f"Could not scrape {tour}: {str(e)}")
                continue
        
        return pd.DataFrame(all_matches) if all_matches else None
        
    except Exception as e:
        st.error(f"Scraping error: {str(e)}")
        return None

@st.cache_data(ttl=3600)
def scrape_espn_tennis():
    """Scrape live matches from ESPN Tennis"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        url = 'https://www.espn.com/tennis/scoreboard'
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        all_matches = []
        
        # Find match containers
        matches = soup.find_all('div', {'class': 'ScoreCell'})
        
        if matches:
            st.success(f"✅ Found {len(matches)} matches on ESPN Tennis")
            
            for match in matches[:30]:
                try:
                    competitors = match.find_all('div', {'class': 'Competitor'})
                    
                    if len(competitors) >= 2:
                        player1_elem = competitors[0].find('a')
                        player2_elem = competitors[1].find('a')
                        
                        if player1_elem and player2_elem:
                            player1 = player1_elem.text.strip()
                            player2 = player2_elem.text.strip()
                            
                            score_elem = match.find('div', {'class': 'Score'})
                            score = score_elem.text.strip() if score_elem else "Live"
                            
                            all_matches.append({
                                'Winner': player1,
                                'Loser': player2,
                                'Score': score,
                                'Tour': 'ATP/WTA',
                                'Surface': 'Hard',
                                'Date': datetime.now().strftime('%Y-%m-%d'),
                                'W1': 0, 'L1': 0, 'W2': 0, 'L2': 0, 'W3': 0, 'L3': 0
                            })
                except:
                    continue
        
        return pd.DataFrame(all_matches) if all_matches else None
        
    except Exception as e:
        st.warning(f"ESPN scraping error: {str(e)}")
        return None

@st.cache_data(ttl=3600)
def load_github_data():
    """Load historical data from GitHub"""
    wta_url = "https://github.com/paulom40/teste/raw/main/wta_data.xlsx"
    atp_url = "https://github.com/paulom40/teste/raw/main/atp_data.xlsx"
    
    dfs = []
    
    try:
        response = requests.get(wta_url, timeout=15)
        response.raise_for_status()
        from io import BytesIO
        wta_df = pd.read_excel(BytesIO(response.content))
        wta_df['Tour'] = 'WTA'
        st.success("✅ WTA historical data loaded")
        dfs.append(wta_df)
    except Exception as e:
        st.warning(f"Could not load WTA: {str(e)}")
    
    try:
        response = requests.get(atp_url, timeout=15)
        response.raise_for_status()
        from io import BytesIO
        atp_df = pd.read_excel(BytesIO(response.content))
        atp_df['Tour'] = 'ATP'
        st.success("✅ ATP historical data loaded")
        dfs.append(atp_df)
    except Exception as e:
        st.warning(f"Could not load ATP: {str(e)}")
    
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    return None

# Load data
st.markdown("### 📊 Loading Data Sources...")

col1, col2 = st.columns(2)

with col1:
    st.markdown("**Live Matches:**")
    live_data = scrape_espn_tennis()
    if live_data is not None:
        st.success(f"✅ {len(live_data)} live matches")
    else:
        st.info("ℹ️ No live matches found")

with col2:
    st.markdown("**Historical Data:**")
    historical_data = load_github_data()
    if historical_data is not None:
        st.success(f"✅ {len(historical_data):,} historical matches")
    else:
        st.warning("⚠️ Could not load historical data")

# Combine data
if historical_data is not None:
    df = historical_data.copy()
else:
    st.error("❌ Could not load any data")
    st.stop()

st.info(f"📊 Total matches: {len(df):,}")
if 'Tour' in df.columns:
    wta_count = len(df[df['Tour'] == 'WTA'])
    atp_count = len(df[df['Tour'] == 'ATP'])
    st.info(f"✅ WTA: {wta_count:,} | ATP: {atp_count:,}")

st.markdown("---")

def calculate_total_games(row):
    """Calculate total games in match"""
    total = 0
    for i in range(1, 6):
        w = row.get(f'W{i}', 0)
        l = row.get(f'L{i}', 0)
        if pd.notna(w) and pd.notna(l) and w > 0 and l > 0:
            total += int(w) + int(l)
    return total if total > 0 else None

def predict_competitiveness(row):
    """Calculate competitiveness score"""
    try:
        w1 = int(row.get('W1', 0)) if pd.notna(row.get('W1')) else 0
        l1 = int(row.get('L1', 0)) if pd.notna(row.get('L1')) else 0
        w2 = int(row.get('W2', 0)) if pd.notna(row.get('W2')) else 0
        l2 = int(row.get('L2', 0)) if pd.notna(row.get('L2')) else 0
        
        set1_closeness = 1 - (abs(w1 - l1) / max(w1 + l1, 1)) if (w1 + l1) > 0 else 0
        set2_closeness = 1 - (abs(w2 - l2) / max(w2 + l2, 1)) if (w2 + l2) > 0 else 0
        
        competitiveness = (set1_closeness + set2_closeness) / 2
        return max(0, min(1, competitiveness))
    except:
        return 0

# Analyze matches
with st.spinner("Analyzing matches..."):
    df_analysis = df.copy()
    df_analysis['Total_Games'] = df_analysis.apply(calculate_total_games, axis=1)
    df_analysis['Competitiveness'] = df_analysis.apply(predict_competitiveness, axis=1)
    df_analysis = df_analysis.dropna(subset=['Total_Games'])

# Filter competitive matches
competitive_matches = df_analysis[
    (df_analysis['Total_Games'] >= 20) & 
    (df_analysis['Total_Games'] <= 26) &
    (df_analysis['Competitiveness'] >= 0.5)
].copy()

competitive_matches = competitive_matches.sort_values('Competitiveness', ascending=False)

st.markdown("---")

# Quick date filters
st.markdown("### ⚡ QUICK DATE FILTERS")
col1, col2, col3 = st.columns(3)

today = datetime.now().date()
tomorrow = today + timedelta(days=1)

with col1:
    if st.button("📅 TODAY ONLY", use_container_width=True, key="today_only"):
        st.session_state.quick_date = today
        st.rerun()

with col2:
    if st.button("📅 TOMORROW ONLY", use_container_width=True, key="tomorrow_only"):
        st.session_state.quick_date = tomorrow
        st.rerun()

with col3:
    if st.button("📅 TODAY + TOMORROW", use_container_width=True, key="today_tomorrow"):
        st.session_state.quick_date = None
        st.rerun()

st.markdown("---")

st.markdown("## 🎯 COMPETITIVE MATCHES")
st.markdown(f"Found **{len(competitive_matches):,} matches** (20-26 games, 50%+ competitiveness)")
st.markdown("---")

# FILTERS
st.markdown("### 🔍 FILTER & ANALYZE")

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    min_games = st.slider("Min Games", 20, 26, 21, key="min_games")

with col2:
    max_games = st.slider("Max Games", 20, 26, 23, key="max_games")

with col3:
    min_competitiveness = st.slider("Min Competitiveness %", 0, 100, 50, 5, key="comp")

with col4:
    tour_filter = st.multiselect("Tours", ['ATP', 'WTA'], default=['ATP', 'WTA'], key="tour")

with col5:
    st.markdown("**📅 Date Range**")
    # Set default to today and tomorrow
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    
    date_range = st.date_input(
        "Select date range",
        value=(today, tomorrow),
        key="date_range"
    )

# Filter matches
filtered_matches = competitive_matches[
    (competitive_matches['Total_Games'] >= min_games) &
    (competitive_matches['Total_Games'] <= max_games) &
    (competitive_matches['Competitiveness'] >= min_competitiveness/100) &
    (competitive_matches['Tour'].isin(tour_filter))
].copy()

# Apply date filter
if 'Date' in filtered_matches.columns:
    try:
        filtered_matches['Date_parsed'] = pd.to_datetime(filtered_matches['Date'], errors='coerce')
        if isinstance(date_range, tuple) and len(date_range) == 2:
            start_date, end_date = date_range
            filtered_matches = filtered_matches[
                (filtered_matches['Date_parsed'].dt.date >= start_date) &
                (filtered_matches['Date_parsed'].dt.date <= end_date)
            ]
            
            # Show date range info
            today = datetime.now().date()
            tomorrow = today + timedelta(days=1)
            
            if start_date == today and end_date == today:
                st.info(f"📅 TODAY ({today.strftime('%d/%m/%Y')})")
            elif start_date == today and end_date == tomorrow:
                st.info(f"📅 TODAY ({today.strftime('%d/%m/%Y')}) + TOMORROW ({tomorrow.strftime('%d/%m/%Y')})")
            elif start_date == tomorrow and end_date == tomorrow:
                st.info(f"📅 TOMORROW ({tomorrow.strftime('%d/%m/%Y')})")
            else:
                st.info(f"📅 {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}")
    except:
        pass

st.markdown(f"### 📊 {len(filtered_matches):,} matches found")

# Display results
if len(filtered_matches) > 0:
    top_n = st.slider("Show top N", 5, 100, min(25, len(filtered_matches)), key="top_n")
    top_matches = filtered_matches.head(top_n)
    
    display_data = []
    for idx, (_, match) in enumerate(top_matches.iterrows(), 1):
        w1 = int(match.get('W1', 0)) if pd.notna(match.get('W1')) else 0
        l1 = int(match.get('L1', 0)) if pd.notna(match.get('L1')) else 0
        w2 = int(match.get('W2', 0)) if pd.notna(match.get('W2')) else 0
        l2 = int(match.get('L2', 0)) if pd.notna(match.get('L2')) else 0
        
        display_data.append({
            '🏆': match['Tour'],
            'Rank': idx,
            'Player 1': match['Winner'],
            'Player 2': match['Loser'],
            'Surface': match.get('Surface', 'N/A'),
            'Score': f"{w1}-{l1} {w2}-{l2}",
            'Games': int(match['Total_Games']),
            'Competitiveness': f"{match['Competitiveness']*100:.1f}%",
            'Date': str(match.get('Date', 'N/A'))
        })
    
    display_df = pd.DataFrame(display_data)
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    st.markdown("## 📥 EXPORT OPTIONS")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Download as Excel", use_container_width=True, key="export_excel"):
            try:
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Create Excel file in memory
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Sheet 1: Summary
                    summary_data = {
                        'Metric': [
                            'Generated',
                            'Total Analyzed',
                            'Competitive Found',
                            'Filtered Results',
                            'Games Range',
                            'Min Competitiveness',
                            'Tours Included',
                            'Date Range Start',
                            'Date Range End'
                        ],
                        'Value': [
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            str(len(df_analysis)),
                            str(len(competitive_matches)),
                            str(len(filtered_matches)),
                            f"{min_games}-{max_games}",
                            f"{min_competitiveness}%",
                            ', '.join(tour_filter),
                            str(date_range[0]) if isinstance(date_range, tuple) else 'N/A',
                            str(date_range[1]) if isinstance(date_range, tuple) else 'N/A'
                        ]
                    }
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Sheet 2: Top Matches
                    top_matches_export = filtered_matches.head(top_n).copy()
                    top_matches_export['Rank'] = range(1, len(top_matches_export) + 1)
                    export_cols = ['Rank', 'Tour', 'Winner', 'Loser', 'Surface', 'W1', 'L1', 'W2', 'L2', 'W3', 'L3', 'Total_Games', 'Competitiveness', 'Date']
                    available_cols = [col for col in export_cols if col in top_matches_export.columns]
                    top_matches_export[available_cols].to_excel(writer, sheet_name='Top Matches', index=False)
                    
                    # Sheet 3: All Filtered Matches
                    all_filtered = filtered_matches.copy()
                    all_filtered['Rank'] = range(1, len(all_filtered) + 1)
                    export_cols_all = ['Rank', 'Tour', 'Winner', 'Loser', 'Surface', 'W1', 'L1', 'W2', 'L2', 'W3', 'L3', 'Total_Games', 'Competitiveness', 'Date']
                    available_cols_all = [col for col in export_cols_all if col in all_filtered.columns]
                    all_filtered[available_cols_all].to_excel(writer, sheet_name='All Filtered', index=False)
                    
                    # Format sheets
                    workbook = writer.book
                    
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        
                        # Header styling
                        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        
                        for cell in worksheet[1]:
                            if cell.value:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Border styling
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                cell.border = thin_border
                                if cell.row > 1:
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                output.seek(0)
                
                st.download_button(
                    label="📊 Download Excel Report",
                    data=output.getvalue(),
                    file_name=f"ATP_WTA_Competitive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success("✅ Excel report ready to download!")
            except ImportError:
                st.error("❌ openpyxl not installed. Using CSV instead.")
                csv = display_df.to_csv(index=False)
                st.download_button(
                    label="📊 Download CSV",
                    data=csv,
                    file_name=f"ATP_WTA_Competitive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"❌ Error creating Excel: {str(e)}")
    
    with col2:
        if st.button("📄 Download as CSV", use_container_width=True, key="export_csv"):
            csv = display_df.to_csv(index=False)
            st.download_button(
                label="📄 Download CSV",
                data=csv,
                file_name=f"ATP_WTA_Competitive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )
            st.success("✅ CSV ready to download!")

else:
    st.warning("No matches found. Try adjusting filters.")

st.markdown("---")
st.markdown("### 📊 STATISTICS")
col1, col2, col3, col4, col5 = st.columns(5)
col1.metric("Total Analyzed", f"{len(df_analysis):,}")
col2.metric("Competitive", f"{len(competitive_matches):,}")
col3.metric("Avg Games", f"{competitive_matches['Total_Games'].mean():.1f}")
col4.metric("Avg Competitiveness", f"{competitive_matches['Competitiveness'].mean()*100:.1f}%")

if 'Tour' in competitive_matches.columns:
    atp_count = len(competitive_matches[competitive_matches['Tour'] == 'ATP'])
    col5.metric("ATP Competitive", f"{atp_count:,}")
